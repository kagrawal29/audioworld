"""Migrate local Excel tracker to Google Sheets.

Usage:
  1. Set up credentials (one of):
     a) Service account: place JSON key at charlie/service_account.json
        and share the sheet with the service account email
     b) OAuth: place client_secret.json at charlie/client_secret.json
        (will open browser for consent on first run)

  2. Run:
     cd /Users/kshitiz/Desktop/Seed Forth/audioworld
     python -m charlie.migrate_to_sheets
"""

import json
import sys
from datetime import datetime
from pathlib import Path

import gspread
import openpyxl

PROJECT_ROOT = Path(__file__).parent.parent
EXCEL_FILE = PROJECT_ROOT / "AudioWorld_March_2026_Touchpoints_with_Content_Library.xlsx"
SERVICE_ACCOUNT_FILE = Path(__file__).parent / "service_account.json"
OAUTH_CREDS_FILE = Path(__file__).parent / "client_secret.json"
TOKEN_FILE = Path(__file__).parent / "token.json"

SHEET_TITLE = "AudioWorld — March 2026 Outreach Tracker"


def get_client() -> gspread.Client:
    """Authenticate with Google Sheets API."""
    if SERVICE_ACCOUNT_FILE.exists():
        return gspread.service_account(filename=str(SERVICE_ACCOUNT_FILE))

    if OAUTH_CREDS_FILE.exists():
        return gspread.oauth(
            credentials_filename=str(OAUTH_CREDS_FILE),
            authorized_user_filename=str(TOKEN_FILE),
        )

    print("No credentials found. Place one of:")
    print(f"  - Service account JSON at: {SERVICE_ACCOUNT_FILE}")
    print(f"  - OAuth client secret at:  {OAUTH_CREDS_FILE}")
    sys.exit(1)


def read_excel() -> dict[str, list[list]]:
    """Read all sheets from Excel file into a dict of sheet_name -> rows."""
    wb = openpyxl.load_workbook(str(EXCEL_FILE), data_only=False)
    sheets = {}
    for name in wb.sheetnames:
        ws = wb[name]
        rows = []
        for row in ws.iter_rows(min_row=1, max_row=ws.max_row, max_col=ws.max_column, values_only=False):
            vals = []
            for cell in row:
                v = cell.value
                if v is None:
                    vals.append("")
                elif isinstance(v, datetime):
                    vals.append(v.strftime("%Y-%m-%d"))
                elif isinstance(v, str) and v.startswith("="):
                    # Preserve formulas
                    vals.append(v)
                else:
                    vals.append(str(v))
            rows.append(vals)
        sheets[name] = rows
    return sheets


def create_and_populate(client: gspread.Client, data: dict[str, list[list]]) -> str:
    """Create Google Sheet and populate all tabs. Returns the spreadsheet URL."""
    spreadsheet = client.create(SHEET_TITLE)
    sheet_id = spreadsheet.id
    print(f"Created spreadsheet: {sheet_id}")

    sheet_names = list(data.keys())

    # Rename the default sheet to the first tab name
    first_ws = spreadsheet.sheet1
    first_ws.update_title(sheet_names[0])
    populate_worksheet(first_ws, data[sheet_names[0]])
    print(f"  Populated: {sheet_names[0]}")

    # Create remaining tabs
    for name in sheet_names[1:]:
        rows = data[name]
        num_rows = max(len(rows), 10)
        num_cols = max(len(rows[0]) if rows else 1, 3)
        ws = spreadsheet.add_worksheet(title=name, rows=num_rows + 5, cols=num_cols + 2)
        populate_worksheet(ws, rows)
        print(f"  Populated: {name}")

    return spreadsheet.url


def populate_worksheet(ws: gspread.Worksheet, rows: list[list]) -> None:
    """Write rows to a worksheet, handling formulas correctly."""
    if not rows:
        return

    num_rows = len(rows)
    num_cols = max(len(r) for r in rows)

    # Pad rows to uniform width
    padded = []
    for row in rows:
        padded.append(row + [""] * (num_cols - len(row)))

    # Resize worksheet if needed
    ws.resize(rows=max(num_rows + 5, ws.row_count), cols=max(num_cols + 2, ws.col_count))

    # Separate values and formulas
    # First pass: write all non-formula values
    value_rows = []
    formula_cells = []
    for i, row in enumerate(padded, start=1):
        value_row = []
        for j, val in enumerate(row):
            if isinstance(val, str) and val.startswith("="):
                value_row.append("")  # placeholder
                formula_cells.append((i, j + 1, val))  # 1-indexed
            else:
                value_row.append(val)
        value_rows.append(value_row)

    # Batch update values
    cell_range = f"A1:{gspread.utils.rowcol_to_a1(num_rows, num_cols)}"
    ws.update(cell_range, value_rows, value_input_option="RAW")

    # Write formulas one by one (Google Sheets API requires USER_ENTERED for formulas)
    if formula_cells:
        batch = []
        for row, col, formula in formula_cells:
            cell_a1 = gspread.utils.rowcol_to_a1(row, col)
            batch.append({"range": cell_a1, "values": [[formula]]})

        # Batch update formulas
        ws.spreadsheet.values_batch_update(
            body={
                "valueInputOption": "USER_ENTERED",
                "data": batch,
            }
        )
        print(f"    Set {len(formula_cells)} formulas")


def set_public_view(client: gspread.Client, spreadsheet_url: str) -> None:
    """Set the spreadsheet to public view-only access."""
    spreadsheet = client.open_by_url(spreadsheet_url)
    spreadsheet.share("", perm_type="anyone", role="reader")
    print("Set public view-only access")


def main():
    print(f"Reading Excel file: {EXCEL_FILE}")
    data = read_excel()
    print(f"Found {len(data)} sheets: {', '.join(data.keys())}")

    print("\nAuthenticating with Google...")
    client = get_client()

    print(f"\nCreating Google Sheet: {SHEET_TITLE}")
    url = create_and_populate(client, data)

    print(f"\nSetting public view access...")
    set_public_view(client, url)

    print(f"\nMigration complete.")
    print(f"Sheet URL: {url}")

    # Save sheet URL for reference
    info = {"sheet_url": url, "migrated_at": datetime.now().isoformat()}
    info_path = Path(__file__).parent / "sheet_info.json"
    info_path.write_text(json.dumps(info, indent=2))
    print(f"Sheet info saved to: {info_path}")

    return url


if __name__ == "__main__":
    main()
